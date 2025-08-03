from flask import Flask, request, jsonify, send_from_directory, render_template_string, send_file
import pandas as pd
import os
import math
import numpy as np
from datetime import datetime
from config import HOST, PORT, VERIFY_BASE_URL

EXCEL_FILE = "vouchers.xlsx"
app = Flask(__name__)

def load_data():
    """加载Excel数据"""
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame(columns=["voucher_id", "date", "customer", "sku", "qty", "price", "status", "used_date"])
    df = pd.read_excel(EXCEL_FILE)
    # 确保新字段存在
    if "used_date" not in df.columns:
        df["used_date"] = ""
    
    # 清理NaN值，将NaN替换为空字符串或适当的默认值
    df = df.fillna({
        "customer": "",
        "sku": "",
        "date": "",
        "used_date": "",
        "status": "未使用",
        "qty": 0,
        "price": 0,
        "voucher_id": ""
    })
    
    # 确保数值类型正确
    df["qty"] = pd.to_numeric(df["qty"], errors='coerce').fillna(0).astype(int)
    df["price"] = pd.to_numeric(df["price"], errors='coerce').fillna(0)
    
    return df

def save_data(df):
    """保存数据到Excel"""
    df.to_excel(EXCEL_FILE, index=False)

# 扫码验证和核销页面
@app.route("/verify/<voucher_id>")
def verify(voucher_id):
    """验证凭证（仅查看，不自动核销）"""
    try:
        df = load_data()
        voucher_row = df[df["voucher_id"] == voucher_id]
        
        if voucher_row.empty:
            return render_template_string("""
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>凭证验证结果</title>
                <style>
                    body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                    .container { max-width: 500px; margin: 0 auto; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
                    .error { background-color: #fff2f0; border: 1px solid #ffccc7; color: #a8071a; }
                    h1 { margin-bottom: 20px; }
                    p { font-size: 16px; margin: 10px 0; }
                </style>
            </head>
            <body>
                <div class="container error">
                    <h1>❌ 凭证无效</h1>
                    <p>凭证号: {{ voucher_id }}</p>
                    <p>该凭证不存在，请检查凭证号是否正确</p>
                </div>
            </body>
            </html>
            """, voucher_id=voucher_id)
        
        voucher = voucher_row.iloc[0]
        status = voucher["status"]
        
        if status == "已使用":
            used_date = voucher.get("used_date", "")
            return render_template_string("""
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>凭证验证结果</title>
                <style>
                    body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                    .container { max-width: 500px; margin: 0 auto; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
                    .warning { background-color: #fffbe6; border: 1px solid #ffe58f; color: #d46b08; }
                    h1 { margin-bottom: 20px; }
                    p { font-size: 16px; margin: 10px 0; }
                    .voucher-info { background: #f5f5f5; padding: 15px; margin: 20px 0; border-radius: 5px; }
                </style>
            </head>
            <body>
                <div class="container warning">
                    <h1>⚠️ 凭证已使用</h1>
                    <div class="voucher-info">
                        <p><strong>凭证号:</strong> {{ voucher_id }}</p>
                        <p><strong>客户:</strong> {{ customer }}</p>
                        <p><strong>商品:</strong> {{ sku }}</p>
                        <p><strong>件数:</strong> {{ qty }}</p>
                        <p><strong>金额:</strong> ¥{{ price }}</p>
                        <p><strong>状态:</strong> {{ status }}</p>
                        {% if used_date %}
                        <p><strong>使用时间:</strong> {{ used_date }}</p>
                        {% endif %}
                    </div>
                    <p>该凭证已经使用过，无法重复使用</p>
                </div>
            </body>
            </html>
            """, 
            voucher_id=voucher_id,
            customer=voucher["customer"],
            sku=voucher["sku"],
            qty=voucher["qty"],
            price=voucher["price"],
            status=voucher["status"],
            used_date=used_date)
        else:
            # 未使用状态：显示凭证信息，但不自动核销
            return render_template_string("""
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>凭证验证结果</title>
                <style>
                    body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                    .container { max-width: 500px; margin: 0 auto; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
                    .success { background-color: #f6ffed; border: 1px solid #b7eb8f; color: #389e0d; }
                    h1 { margin-bottom: 20px; }
                    p { font-size: 16px; margin: 10px 0; }
                    .voucher-info { background: #f5f5f5; padding: 15px; margin: 20px 0; border-radius: 5px; }
                    .notice { background: #e6f7ff; border: 1px solid #91d5ff; color: #0050b3; padding: 15px; margin: 20px 0; border-radius: 5px; }
                </style>
            </head>
            <body>
                <div class="container success">
                    <h1>✅ 凭证有效</h1>
                    <div class="voucher-info">
                        <p><strong>凭证号:</strong> {{ voucher_id }}</p>
                        <p><strong>客户:</strong> {{ customer }}</p>
                        <p><strong>商品:</strong> {{ sku }}</p>
                        <p><strong>件数:</strong> {{ qty }}</p>
                        <p><strong>金额:</strong> ¥{{ price }}</p>
                        <p><strong>状态:</strong> {{ status }}</p>
                        <p><strong>创建日期:</strong> {{ date }}</p>
                    </div>
                    <div class="notice">
                        <p><strong>注意:</strong> 此凭证有效且未使用</p>
                        <p>如需核销，请在管理系统中操作</p>
                    </div>
                </div>
            </body>
            </html>
            """, 
            voucher_id=voucher_id,
            customer=voucher["customer"],
            sku=voucher["sku"],
            qty=voucher["qty"],
            price=voucher["price"],
            status=voucher["status"],
            date=voucher["date"])
        
    except Exception as e:
        print(f"验证凭证错误: {e}")
        return render_template_string("""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>验证错误</title>
            <style>
                body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                .container { max-width: 500px; margin: 0 auto; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
                .error { background-color: #fff2f0; border: 1px solid #ffccc7; color: #a8071a; }
                h1 { margin-bottom: 20px; }
                p { font-size: 16px; margin: 10px 0; }
            </style>
        </head>
        <body>
            <div class="container error">
                <h1>❌ 系统错误</h1>
                <p>验证过程中发生错误，请稍后再试</p>
            </div>
        </body>
        </html>
        """), 500

# 静态文件路由
@app.route("/static/<path:filename>")
def static_files(filename):
    """提供静态文件"""
    return send_from_directory("static", filename)

# 管理后台主页
@app.route("/")
def index():
    """管理后台主页"""
    return send_from_directory(".", "index.html")

# Web管理API - 获取所有凭证
@app.route("/api/list")
def list_vouchers():
    """获取所有凭证列表"""
    try:
        df = load_data()
        
        # 转换为字典列表
        records = df.to_dict(orient="records")
        
        # 额外清理：确保没有NaN、inf等不合法的JSON值
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if pd.isna(value) or value is None:
                    # 根据字段类型设置默认值
                    if key in ["qty", "price"]:
                        cleaned_record[key] = 0
                    else:
                        cleaned_record[key] = ""
                elif isinstance(value, float):
                    # 检查是否为无穷大或NaN
                    if pd.isna(value) or not np.isfinite(value):
                        if key in ["qty", "price"]:
                            cleaned_record[key] = 0
                        else:
                            cleaned_record[key] = ""
                    else:
                        cleaned_record[key] = value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)
        
        return jsonify(cleaned_records)
        
    except Exception as e:
        print(f"获取凭证列表错误: {e}")
        return jsonify({"error": str(e)}), 500

# Web管理API - 添加凭证
@app.route("/api/add", methods=["POST"])
def add_voucher():
    """添加新凭证"""
    try:
        data = request.json
        if not data:
            return jsonify({"error": "请求数据为空"}), 400
        
        df = load_data()
        
        # 验证必需字段
        required_fields = ["voucher_id", "customer", "sku", "qty", "price"]
        for field in required_fields:
            if field not in data or data[field] is None or data[field] == "":
                return jsonify({"error": f"缺少必需字段: {field}"}), 400
        
        # 数据类型验证和转换
        try:
            data["qty"] = int(data["qty"])
            data["price"] = float(data["price"])
        except (ValueError, TypeError):
            return jsonify({"error": "件数和价格必须是有效数字"}), 400
        
        # 数值范围验证
        if data["qty"] <= 0:
            return jsonify({"error": "件数必须大于0"}), 400
        if data["price"] < 0:
            return jsonify({"error": "价格不能小于0"}), 400
        
        # 设置默认值
        if "date" not in data or not data["date"]:
            data["date"] = datetime.now().strftime("%Y-%m-%d")
        if "status" not in data:
            data["status"] = "未使用"
        if "used_date" not in data:
            data["used_date"] = ""
        
        # 检查凭证号是否已存在
        if not df[df["voucher_id"] == data["voucher_id"]].empty:
            return jsonify({"error": "凭证号已存在"}), 400
        
        # 验证字符串长度
        if len(data["customer"]) > 50:
            return jsonify({"error": "客户名称过长"}), 400
        if len(data["sku"]) > 50:
            return jsonify({"error": "款号过长"}), 400
        
        # 确保所有必需字段都有值，避免NaN
        clean_data = {
            "voucher_id": str(data["voucher_id"]).strip(),
            "date": str(data["date"]).strip(),
            "customer": str(data["customer"]).strip(),
            "sku": str(data["sku"]).strip(),
            "qty": int(data["qty"]),
            "price": float(data["price"]),
            "status": str(data.get("status", "未使用")).strip(),
            "used_date": str(data.get("used_date", "")).strip()
        }
        
        # 将新数据添加到最前面
        new_row = pd.DataFrame([clean_data])
        df = pd.concat([new_row, df], ignore_index=True)
        save_data(df)
        return jsonify({"msg": "添加成功", "voucher_id": data["voucher_id"]})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Web管理API - 更新凭证
@app.route("/api/update/<voucher_id>", methods=["POST"])
def update_voucher(voucher_id):
    """更新凭证信息"""
    try:
        data = request.json
        if not data:
            return jsonify({"error": "请求数据为空"}), 400
        
        df = load_data()
        
        if df[df["voucher_id"] == voucher_id].empty:
            return jsonify({"error": "凭证不存在"}), 404
        
        # 验证和转换数据类型
        for key, value in data.items():
            if key == "qty":
                try:
                    data[key] = int(value)
                    if data[key] <= 0:
                        return jsonify({"error": "件数必须大于0"}), 400
                except (ValueError, TypeError):
                    return jsonify({"error": "件数必须是有效整数"}), 400
            elif key == "price":
                try:
                    data[key] = float(value)
                    if data[key] < 0:
                        return jsonify({"error": "价格不能小于0"}), 400
                except (ValueError, TypeError):
                    return jsonify({"error": "价格必须是有效数字"}), 400
            elif key in ["customer", "sku"]:
                # 确保转换为字符串后再处理
                str_value = str(value) if value is not None else ""
                if not str_value or str_value.strip() == "":
                    return jsonify({"error": f"{key}不能为空"}), 400
                if len(str_value.strip()) > 50:
                    return jsonify({"error": f"{key}过长"}), 400
                data[key] = str_value.strip()
            elif key == "status":
                if value not in ["未使用", "已使用"]:
                    return jsonify({"error": "状态值无效"}), 400
        
        # 更新数据
        for k, v in data.items():
            df.loc[df["voucher_id"] == voucher_id, k] = v
            
        save_data(df)
        return jsonify({"msg": "更新成功"})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Web管理API - 删除凭证
@app.route("/api/delete/<voucher_id>", methods=["POST"])
def delete_voucher(voucher_id):
    """删除凭证"""
    try:
        df = load_data()
        
        if df[df["voucher_id"] == voucher_id].empty:
            return jsonify({"error": "凭证不存在"}), 404
        
        df = df[df["voucher_id"] != voucher_id]
        save_data(df)
        return jsonify({"msg": "删除成功"})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 生成票据图片API
@app.route("/api/generate_voucher/<voucher_id>", methods=["POST"])
def generate_voucher_image(voucher_id):
    """为现有凭证生成票据图片和二维码"""
    try:
        df = load_data()
        voucher = df[df["voucher_id"] == voucher_id]
        
        if voucher.empty:
            return jsonify({"error": "凭证不存在"}), 404
        
        voucher_info = voucher.iloc[0].to_dict()
        
        # 导入生成模块
        import qrcode
        from PIL import Image, ImageDraw, ImageFont
        import os
        
        QRCODE_DIR = "vouchers"
        BASE_URL = VERIFY_BASE_URL
        
        # 确保目录存在
        os.makedirs(QRCODE_DIR, exist_ok=True)
        
        # 生成二维码
        url = BASE_URL + voucher_id
        qrcode_path = f"{QRCODE_DIR}/{voucher_id}_qr.png"
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_img.save(qrcode_path)
        
        # 创建票据图片
        try:
            qr_img = Image.open(qrcode_path).resize((200, 200))
            img = Image.new('RGB', (500, 300), 'white')
            draw = ImageDraw.Draw(img)
            
            # 字体处理
            try:
                # macOS中文字体
                font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 22)
            except:
                try:
                    # 其他中文字体
                    font = ImageFont.truetype("/System/Library/Fonts/STHeiti Light.ttc", 22)
                except:
                    try:
                        # Windows中文字体
                        font = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 22)
                    except:
                        try:
                            # Linux中文字体
                            font = ImageFont.truetype("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", 22)
                        except:
                            # 默认字体
                            font = ImageFont.load_default()
            
            # 绘制信息
            draw.text((20, 20), "商家：金潮男装供应链", fill='black', font=font)
            draw.text((20, 50), f"客户：{voucher_info['customer']}", fill='black', font=font)
            draw.text((20, 80), f"款号：{voucher_info['sku']}", fill='black', font=font)
            draw.text((20, 110), f"件数：{voucher_info['qty']}", fill='black', font=font)
            draw.text((20, 140), f"金额：{voucher_info['price']} 元", fill='black', font=font)
            draw.text((20, 170), f"日期：{voucher_info['date']}", fill='black', font=font)
            draw.text((20, 200), f"凭证号：{voucher_info['voucher_id']}", fill='black', font=font)
            
            # 添加边框
            draw.rectangle([(10, 10), (490, 290)], outline='black', width=2)
            
            # 粘贴二维码
            img.paste(qr_img, (280, 50))
            
            # 保存票据图片
            voucher_path = f"{QRCODE_DIR}/{voucher_id}.png"
            img.save(voucher_path)
            
            return jsonify({
                "msg": "票据生成成功",
                "voucher_image": f"/vouchers/{voucher_id}.png",
                "qrcode_image": f"/vouchers/{voucher_id}_qr.png",
                "verify_url": url
            })
            
        except Exception as e:
            return jsonify({"error": f"生成票据图片失败: {str(e)}"}), 500
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 提供静态文件服务（票据图片和二维码）
@app.route("/vouchers/<path:filename>")
def serve_vouchers(filename):
    """提供票据文件访问"""
    return send_from_directory("vouchers", filename)

# 健康检查接口
@app.route("/health")
def health():
    """健康检查"""
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()})

# 新增：手动核销凭证API
@app.route("/api/use_voucher/<voucher_id>", methods=["POST"])
def use_voucher(voucher_id):
    """手动核销凭证"""
    try:
        df = load_data()
        
        # 检查凭证是否存在
        voucher_row = df[df["voucher_id"] == voucher_id]
        if voucher_row.empty:
            return jsonify({"error": "凭证不存在"}), 404
        
        # 检查凭证状态
        if voucher_row.iloc[0]["status"] == "已使用":
            return jsonify({"error": "凭证已经使用过"}), 400
        
        # 更新状态为已使用
        df.loc[df["voucher_id"] == voucher_id, "status"] = "已使用"
        df.loc[df["voucher_id"] == voucher_id, "used_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 保存数据
        save_data(df)
        
        return jsonify({
            "msg": "凭证核销成功",
            "voucher_id": voucher_id,
            "used_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
    except Exception as e:
        print(f"核销凭证错误: {e}")
        return jsonify({"error": str(e)}), 500

# Excel导出API
@app.route("/api/export", methods=["GET"])
def export_excel():
    """导出Excel文件"""
    try:
        print("=== 开始导出Excel ===")
        df = load_data()
        print(f"加载数据: {len(df)} 行")
        
        if df.empty:
            print("错误: 没有数据可以导出")
            return jsonify({"error": "没有数据可以导出"}), 400
        
        # 生成导出文件名（包含时间戳）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_filename = f"vouchers_export_{timestamp}.xlsx"
        export_path = os.path.join(os.getcwd(), export_filename)
        print(f"导出文件路径: {export_path}")
        
        # 确保目录存在并有写入权限
        try:
            # 测试写入权限
            test_file = os.path.join(os.getcwd(), f"test_write_{timestamp}.tmp")
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            print("写入权限检查: 通过")
        except Exception as e:
            print(f"写入权限检查失败: {e}")
            return jsonify({"error": f"目录写入权限不足: {str(e)}"}), 500
        
        # 数据预处理 - 确保数据类型正确
        try:
            print("开始数据预处理...")
            print(f"原始数据结构: {df.dtypes}")
            print(f"数据样本:\n{df.head()}")
            
            # 确保所有列都存在且数据类型正确
            required_columns = ["voucher_id", "date", "customer", "sku", "qty", "price", "status", "used_date"]
            for col in required_columns:
                if col not in df.columns:
                    print(f"添加缺失列: {col}")
                    df[col] = ""
            
            # 数据类型转换和清理
            df = df.copy()  # 避免修改原数据
            df['voucher_id'] = df['voucher_id'].astype(str)
            df['customer'] = df['customer'].astype(str) 
            df['sku'] = df['sku'].astype(str)
            df['date'] = df['date'].astype(str)
            df['status'] = df['status'].astype(str)
            df['used_date'] = df['used_date'].astype(str)
            
            # 数值类型处理
            df['qty'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0).astype(int)
            df['price'] = pd.to_numeric(df['price'], errors='coerce').fillna(0.0)
            
            # 替换NaN值
            df = df.fillna('')
            
            print(f"预处理后数据结构: {df.dtypes}")
            print(f"预处理后数据样本:\n{df.head()}")
            
        except Exception as e:
            print(f"数据预处理失败: {e}")
            return jsonify({"error": f"数据预处理失败: {str(e)}"}), 500
        
        # 导出到临时文件
        try:
            print("开始写入Excel文件...")
            
            # 尝试不同的写入方式
            try:
                # 方法1: 使用openpyxl引擎
                df.to_excel(export_path, index=False, engine='openpyxl')
                print("使用openpyxl引擎写入成功")
            except Exception as openpyxl_error:
                print(f"openpyxl引擎失败: {openpyxl_error}")
                try:
                    # 方法2: 使用xlsxwriter引擎（如果可用）
                    df.to_excel(export_path, index=False, engine='xlsxwriter')
                    print("使用xlsxwriter引擎写入成功")
                except Exception as xlsxwriter_error:
                    print(f"xlsxwriter引擎也失败: {xlsxwriter_error}")
                    # 方法3: 手动创建Excel文件
                    import openpyxl
                    from openpyxl import Workbook
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Vouchers"
                    
                    # 写入标题行
                    headers = list(df.columns)
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=str(header))
                    
                    # 写入数据行
                    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                        for col_idx, value in enumerate(row, 1):
                            # 确保值不是NaN或None
                            if pd.isna(value) or value is None:
                                cell_value = ""
                            else:
                                cell_value = str(value) if not isinstance(value, (int, float)) else value
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    wb.save(export_path)
                    print("手动创建Excel文件成功")
            
            # 验证文件
            if os.path.exists(export_path):
                file_size = os.path.getsize(export_path)
                print(f"Excel文件写入成功，文件大小: {file_size} bytes")
                
                if file_size == 0:
                    print("错误: 生成的文件为空")
                    return jsonify({"error": "生成的Excel文件为空"}), 500
                    
            else:
                print("错误: Excel文件未生成")
                return jsonify({"error": "Excel文件生成失败"}), 500
                
        except Exception as e:
            print(f"Excel写入失败: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Excel文件生成失败: {str(e)}"}), 500
        
        # 验证文件是否存在
        if not os.path.exists(export_path):
            print("错误: Excel文件未生成")
            return jsonify({"error": "Excel文件生成失败"}), 500
        
        # 发送文件
        try:
            print("开始发送文件...")
            response = send_file(
                export_path, 
                as_attachment=True, 
                download_name=export_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # 设置响应后删除文件的回调
            def cleanup_file():
                try:
                    if os.path.exists(export_path):
                        os.remove(export_path)
                        print(f"清理临时文件: {export_path}")
                except Exception as cleanup_error:
                    print(f"清理文件失败: {cleanup_error}")
            
            response.call_on_close(cleanup_file)
            print("文件发送成功")
            return response
            
        except Exception as e:
            print(f"文件发送失败: {e}")
            # 清理失败的文件
            try:
                if os.path.exists(export_path):
                    os.remove(export_path)
            except:
                pass
            return jsonify({"error": f"文件发送失败: {str(e)}"}), 500
        
    except Exception as e:
        print(f"导出异常: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"导出失败: {str(e)}"}), 500

# Excel导入API
@app.route("/api/import", methods=["POST"])
def import_excel():
    """导入Excel文件（追加到最前面）"""
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            return jsonify({"error": "没有选择文件"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "没有选择文件"}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "请选择Excel文件（.xlsx或.xls格式）"}), 400
        
        # 读取上传的Excel文件
        try:
            import_df = pd.read_excel(file)
        except Exception as e:
            return jsonify({"error": f"Excel文件读取失败: {str(e)}"}), 400
        
        if import_df.empty:
            return jsonify({"error": "Excel文件为空"}), 400
        
        # 验证必需的列
        required_columns = ["voucher_id", "customer", "sku", "qty", "price"]
        missing_columns = [col for col in required_columns if col not in import_df.columns]
        if missing_columns:
            return jsonify({"error": f"Excel文件缺少必需的列: {', '.join(missing_columns)}"}), 400
        
        # 清理和验证导入数据
        valid_rows = []
        error_rows = []
        
        for index, row in import_df.iterrows():
            try:
                # 验证必需字段
                if pd.isna(row['voucher_id']) or str(row['voucher_id']).strip() == '':
                    error_rows.append(f"第{index+2}行: 凭证号不能为空")
                    continue
                if pd.isna(row['customer']) or str(row['customer']).strip() == '':
                    error_rows.append(f"第{index+2}行: 客户不能为空")
                    continue
                if pd.isna(row['sku']) or str(row['sku']).strip() == '':
                    error_rows.append(f"第{index+2}行: 款号不能为空")
                    continue
                
                # 验证和转换数值
                try:
                    qty = int(row['qty']) if not pd.isna(row['qty']) else 1
                    price = float(row['price']) if not pd.isna(row['price']) else 0.0
                except (ValueError, TypeError):
                    error_rows.append(f"第{index+2}行: 件数和价格必须是数字")
                    continue
                
                if qty <= 0:
                    error_rows.append(f"第{index+2}行: 件数必须大于0")
                    continue
                if price < 0:
                    error_rows.append(f"第{index+2}行: 价格不能小于0")
                    continue
                
                # 构造清理后的数据
                clean_row = {
                    "voucher_id": str(row['voucher_id']).strip(),
                    "date": str(row.get('date', datetime.now().strftime("%Y-%m-%d"))).strip(),
                    "customer": str(row['customer']).strip(),
                    "sku": str(row['sku']).strip(),
                    "qty": qty,
                    "price": price,
                    "status": str(row.get('status', '未使用')).strip(),
                    "used_date": str(row.get('used_date', '')).strip()
                }
                
                valid_rows.append(clean_row)
                
            except Exception as e:
                error_rows.append(f"第{index+2}行: 数据处理错误 - {str(e)}")
        
        if not valid_rows:
            error_msg = "没有有效的数据行"
            if error_rows:
                error_msg += f"\n错误信息:\n" + "\n".join(error_rows[:10])  # 只显示前10个错误
            return jsonify({"error": error_msg}), 400
        
        # 加载现有数据
        existing_df = load_data()
        
        # 检查重复的凭证号
        import_df_clean = pd.DataFrame(valid_rows)
        existing_voucher_ids = set(existing_df['voucher_id'].values) if not existing_df.empty else set()
        duplicate_ids = []
        
        for _, row in import_df_clean.iterrows():
            if row['voucher_id'] in existing_voucher_ids:
                duplicate_ids.append(row['voucher_id'])
        
        if duplicate_ids:
            return jsonify({
                "error": f"以下凭证号已存在，请修改后重新导入: {', '.join(duplicate_ids[:10])}" + 
                        (f" (还有{len(duplicate_ids)-10}个重复)" if len(duplicate_ids) > 10 else "")
            }), 400
        
        # 将导入的数据添加到最前面
        if existing_df.empty:
            final_df = import_df_clean
        else:
            final_df = pd.concat([import_df_clean, existing_df], ignore_index=True)
        
        save_data(final_df)
        
        result = {
            "msg": "导入成功",
            "imported_count": len(valid_rows),
            "total_count": len(final_df)
        }
        
        if error_rows:
            result["warnings"] = f"有{len(error_rows)}行数据存在问题，已跳过"
            result["error_details"] = error_rows[:10]  # 只返回前10个错误详情
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": f"导入失败: {str(e)}"}), 500

# Excel模板下载API
@app.route("/api/template", methods=["GET"])
def download_template():
    """下载Excel导入模板"""
    try:
        print("=== 生成Excel模板 ===")
        
        # 创建模板数据
        template_data = {
            "voucher_id": ["V001", "V002", "V003"],
            "customer": ["张三", "李四", "王五"],
            "sku": ["SKU001", "SKU002", "SKU003"],
            "qty": [1, 2, 1],
            "price": [100.00, 200.50, 150.00],
            "date": ["2024-01-01", "2024-01-02", "2024-01-03"],
            "status": ["未使用", "未使用", "未使用"],
            "used_date": ["", "", ""]
        }
        
        df = pd.DataFrame(template_data)
        
        # 生成模板文件名
        template_filename = "voucher_import_template.xlsx"
        template_path = os.path.join(os.getcwd(), template_filename)
        
        # 导出模板
        df.to_excel(template_path, index=False, engine='openpyxl')
        print(f"模板文件创建成功: {template_path}")
        
        # 发送模板文件
        response = send_file(
            template_path,
            as_attachment=True,
            download_name=template_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 设置清理回调
        def cleanup_template():
            try:
                if os.path.exists(template_path):
                    os.remove(template_path)
                    print(f"清理模板文件: {template_path}")
            except Exception as e:
                print(f"清理模板文件失败: {e}")
        
        response.call_on_close(cleanup_template)
        print("模板文件发送成功")
        return response
        
    except Exception as e:
        print(f"生成模板失败: {e}")
        return jsonify({"error": f"生成模板失败: {str(e)}"}), 500


if __name__ == "__main__":
    try:
        print("=== VoucherLite 服务器启动中 ===")
        print(f"管理后台: http://{HOST}:{PORT}")
        print(f"健康检查: http://{HOST}:{PORT}/health")
        print("API文档:")
        print("- GET  /api/list - 获取所有凭证")
        print("- POST /api/add - 添加凭证")
        print("- POST /api/update/<id> - 更新凭证")
        print("- POST /api/delete/<id> - 删除凭证")
        print("- POST /api/use_voucher/<id> - 核销凭证")
        print("- GET  /verify/<id> - 验证凭证")
        print("- GET  /api/export - 导出Excel文件")
        print("- GET  /api/template - 下载Excel导入模板")
        print("- POST /api/import - 导入Excel文件")
        print("按 Ctrl+C 停止服务器")
        print("-" * 50)
        
        app.run(host="0.0.0.0", port=PORT, debug=False)
        
    except KeyboardInterrupt:
        print("\n服务器已停止")
    except Exception as e:
        print(f"\n❌ 服务器启动失败: {e}")
        import traceback
        traceback.print_exc()
        input("按回车键退出...")
    except SystemExit:
        print("\n程序退出")